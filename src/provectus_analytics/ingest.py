"""Load FSP exports into the SQLite DB.

Two ingest paths:
    Synthetic (tests):  CSV files via ingest_all()
    Real FSP data:      Flight Detail XLSX via ingest_flight_detail_xlsx()

After ingest:
    - students table has one row per FSP client (no survey link yet)
    - flights / invoices loaded (student_id still NULL — reconcile.py sets it)
    - surveys raw responses stored

Reconciliation, enrollment building, partitioning, and milestone computation
all happen downstream.
"""
from __future__ import annotations

import csv
import json
import sqlite3
from pathlib import Path

# ── Aircraft class lookup ─────────────────────────────────────────────────────
# Keyed by FSP model string.  Add real tail/model combos as they appear.
AIRCRAFT_CLASS = {
    # Synthetic fleet
    "172N": "SE_BASIC", "172S": "SE_BASIC", "PA-28-181": "SE_BASIC",
    "182RG": "SE_COMPLEX", "PA-28R-201": "SE_COMPLEX",
    "PA-44-180": "ME_BASIC",   # Piper Seminole (multi-engine)
    # Real Provectus fleet (extend as needed)
    "PA-28-180": "SE_BASIC",   # Piper Cherokee 180
    "PA-28-160": "SE_BASIC",   # Piper Cherokee 160
    "C172M": "SE_BASIC",       # Cessna 172M
    "PA-28R-180": "SE_COMPLEX",
}

# ── Billing column sets ───────────────────────────────────────────────────────
_PRIMARY_COLS = {
    "PRIMARY Hrs", "PRIMARY F&F Hrs",
    "PRIMARY OLD - 85 Hrs", "PRIMARY FIRST RESPONDERS Hrs",
}
_MISC_COLS = {"ADVANCED Hrs", "ATP Hrs"}


def classify_aircraft(model: str | None) -> str | None:
    if not model:
        return None
    return AIRCRAFT_CLASS.get(model.strip(), "SE_BASIC")


def parse_aircraft_string(s: str | None) -> tuple[str | None, str | None, str | None]:
    """'N512PT Piper PA-28-180' → (tail, make, model). Handles null/blank."""
    if not s or not s.strip():
        return None, None, None
    parts = s.strip().split(None, 2)
    if len(parts) == 1:
        return parts[0], None, None
    if len(parts) == 2:
        return parts[0], parts[1], None
    return parts[0], parts[1], parts[2]


def derive_billing_category(row: dict) -> str:
    """Derive a single billing category from a Flight Detail row dict.

    Priority: AMEL > MEI > CFI > CFII > PRIMARY (any variant) > MISC > NONE.
    NONE is returned for Student Solos, Check Rides, and ground lessons with
    no instructor billing — those are attributed via date windows downstream.
    """
    def hrs(col: str) -> float:
        v = row.get(col)
        try:
            return float(v) if v else 0.0
        except (TypeError, ValueError):
            return 0.0

    if hrs("MULTI ENGINE Hrs") > 0:
        return "AMEL"
    if hrs("MEI Hrs") > 0:
        return "MEI"
    if hrs("CFI-A Hrs") > 0:
        return "CFI"
    if hrs("CFI-I Hrs") > 0:
        return "CFII"
    if any(hrs(c) > 0 for c in _PRIMARY_COLS):
        return "PRIMARY"
    if any(hrs(c) > 0 for c in _MISC_COLS):
        return "MISC"
    return "NONE"


# ── Synthetic CSV ingest (kept for tests + synthetic pipeline) ────────────────

def ingest_clients(conn: sqlite3.Connection, clients_csv: Path) -> int:
    """Load FSP client roster into students table."""
    with open(clients_csv) as f:
        rows = list(csv.DictReader(f))
    conn.executemany(
        """INSERT INTO students (fsp_client_id, fsp_display_name, email, match_status)
           VALUES (?, ?, ?, 'unmatched')""",
        [(r["Client ID"], r["Display Name"], r["Email"]) for r in rows],
    )
    conn.commit()
    return len(rows)


def ingest_reservations(conn: sqlite3.Connection, reservations_csv: Path) -> int:
    """Load synthetic FSP reservations CSV → flights table.
    Uses length_hrs; hobbs_hours/billing_category left NULL (synthetic fallback).
    """
    with open(reservations_csv) as f:
        rows = list(csv.DictReader(f))
    conn.executemany(
        """INSERT INTO flights (
               fsp_reservation, fsp_flight_num, flight_date, length_hrs,
               reservation_type, status, client_raw, aircraft_tail, aircraft_make,
               aircraft_model, aircraft_class, instructor
           ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        [
            (
                r["Reservation #"],
                r["Flight #"] or None,
                r["Date"],
                float(r["Length (hrs)"]) if r["Length (hrs)"] else 0.0,
                r["Reservation Type"],
                r["Status"],
                r["Client"] or None,
                r["Aircraft Tail"] or None,
                r["Aircraft Make"] or None,
                r["Aircraft Model"] or None,
                classify_aircraft(r["Aircraft Model"]),
                r["Instructor"] or None,
            )
            for r in rows
        ],
    )
    conn.commit()
    return len(rows)


def ingest_invoices(conn: sqlite3.Connection, invoices_csv: Path) -> int:
    """Load FSP invoice lines. flight_id resolved by reservation number lookup."""
    with open(invoices_csv) as f:
        rows = list(csv.DictReader(f))
    res_map = {
        row["fsp_reservation"]: row["flight_id"]
        for row in conn.execute("SELECT flight_id, fsp_reservation FROM flights")
    }
    conn.executemany(
        """INSERT INTO invoices (
               fsp_invoice, invoice_date, flight_id, description, category,
               qty, rate, amount, status
           ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        [
            (
                r["Invoice #"],
                r["Invoice Date"],
                res_map.get(r["Reservation #"]),
                r["Line Item Description"],
                r["Category"],
                float(r["Quantity (hrs/units)"]) if r["Quantity (hrs/units)"] else None,
                float(r["Rate ($)"]) if r["Rate ($)"] else None,
                float(r["Amount ($)"]),
                r["Status"],
            )
            for r in rows
        ],
    )
    conn.commit()
    return len(rows)


def ingest_survey(conn: sqlite3.Connection, survey_csv: Path) -> int:
    """Store raw survey responses. Linking to students happens in reconcile.py."""
    with open(survey_csv) as f:
        rows = list(csv.DictReader(f))
    conn.executemany(
        """INSERT INTO surveys (student_id, submitted_at, raw_response)
           VALUES (NULL, ?, ?)""",
        [(r.get("Timestamp"), json.dumps(r)) for r in rows],
    )
    conn.commit()
    return len(rows)


# ── Stage-check ingest ────────────────────────────────────────────────────────
# Normalize the form's human-readable labels into stable codes. Keys are matched
# case-insensitively by prefix so minor wording tweaks on the form still land.
_STAGE_LABELS = [
    ("pre-solo", "pre_solo"),
    ("pre-cross", "pre_xc"),
    ("stage 1", "xc_complete"),
    ("xc time", "xc_complete"),
    ("stage 2", "pre_checkride"),
    ("pre-checkride", "pre_checkride"),
]
_RESULT_LABELS = [
    ("satisfactory", "satisfactory"),
    ("unsatisfactory", "unsatisfactory"),
    ("partial", "partial"),
]
# Aircraft family → engine class. The two "Other" buckets keep the SE/ME signal.
_AIRCRAFT_ENGINE = {
    "C172": "SE", "PA-28": "SE", "BE-76": "ME",
    "Other single-engine (ASEL)": "SE", "Other multi-engine (AMEL)": "ME",
}


def _norm_rating(label: str | None) -> str:
    """Form uses 'ASEL COM'; the DB rating code is 'COM'. Others pass through."""
    s = (label or "").strip().upper()
    return "COM" if s in ("ASEL COM", "ASEL COMM", "COMMERCIAL") else s


def _norm_stage(label: str | None) -> str:
    s = (label or "").strip().lower()
    for needle, code in _STAGE_LABELS:
        if needle in s:
            return code
    return s or "unknown"


def _norm_result(label: str | None) -> str:
    s = (label or "").strip().lower()
    for needle, code in _RESULT_LABELS:
        if needle in s:
            return code
    return s or "unknown"


def _engine_class(aircraft: str | None) -> str | None:
    if not aircraft:
        return None
    a = aircraft.strip()
    if a in _AIRCRAFT_ENGINE:
        return _AIRCRAFT_ENGINE[a]
    low = a.lower()
    if "multi" in low or "amel" in low:
        return "ME"
    if "single" in low or "asel" in low:
        return "SE"
    return None


def ingest_stage_checks(conn: sqlite3.Connection, stage_csv: Path) -> int:
    """Load stage-check form responses into `stage_checks` (student_id NULL).

    Reconciliation to a student happens in reconcile.reconcile_stage_checks().
    The CSV is the canonical/normalized shape (see Stage Check Form Spec); the
    real Google-Form response sheet gets mapped to these columns at export time.
    """
    with open(stage_csv, newline="") as f:
        rows = list(csv.DictReader(f))
    for r in rows:
        conn.execute(
            """INSERT INTO stage_checks
                 (student_id, student_name, student_email, check_date, rating, stage,
                  result, hobbs_hours, conducting_instructor, primary_cfi, aircraft,
                  engine_class, notes, match_status, raw_json)
               VALUES (NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'unmatched', ?)""",
            (
                (r.get("student_name") or "").strip(),
                (r.get("student_email") or "").strip(),
                (r.get("date") or "").strip(),
                _norm_rating(r.get("rating")),
                _norm_stage(r.get("stage")),
                _norm_result(r.get("result")),
                _float(r.get("hobbs_hours")),
                (r.get("conducting_instructor") or "").strip() or None,
                (r.get("primary_cfi") or "").strip() or None,
                (r.get("aircraft") or "").strip() or None,
                _engine_class(r.get("aircraft")),
                (r.get("notes") or "").strip() or None,
                json.dumps(r),
            ),
        )
    conn.commit()
    return len(rows)


# Columns in the alumni survey XLSX that hold rating-boundary dates (datetime
# in Excel, but partition.py expects "Month YYYY" strings — same format the
# synthetic CSV uses).
_SURVEY_MONTH_COLUMNS = {
    "PPL training start", "First solo", "XC solos complete", "PPL checkride",
    "IFR training start", "XC PIC complete", "IFR checkride",
    "ASEL COM start", "ASEL COM checkride",
    "AMEL start", "AMEL checkride",
    "CFI start", "CFI checkride",
    "CFII start", "CFII checkride",
    "MEI start", "MEI checkride",
}


def _normalize_survey_cell(col: str, value):
    """Coerce one XLSX cell into the string shape downstream consumers expect.

    - Timestamp     → 'M/D/YYYY H:MM:SS' (matches synthetic CSV)
    - Date columns  → 'Month YYYY'        (matches synthetic CSV → partition.py)
    - Anything else → str() or ''         (Yes/No/freeform passthrough)
    """
    if value is None:
        return ""
    if col == "Timestamp" and hasattr(value, "strftime"):
        return value.strftime("%-m/%-d/%Y %-H:%M:%S")
    if col in _SURVEY_MONTH_COLUMNS and hasattr(value, "strftime"):
        return value.strftime("%B %Y")
    return str(value).strip()


def ingest_survey_xlsx(conn: sqlite3.Connection, xlsx_path: Path) -> int:
    """Ingest a Google-Form-exported alumni survey XLSX.

    The real survey has identical column names to the synthetic CSV, but
    Excel stores dates as datetime objects. Normalize to strings so the
    existing reconcile + partition logic works unchanged.
    """
    wb = _load_workbook(xlsx_path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in
               next(ws.iter_rows(min_row=1, max_row=1))]

    rows_to_insert = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        normalized = {h: _normalize_survey_cell(h, v) for h, v in zip(headers, row)}
        # Require a name OR an email to count as a real response.
        if not normalized.get("Full name") and not normalized.get("Email"):
            continue
        rows_to_insert.append((normalized.get("Timestamp"), json.dumps(normalized)))

    conn.executemany(
        """INSERT INTO surveys (student_id, submitted_at, raw_response)
           VALUES (NULL, ?, ?)""",
        rows_to_insert,
    )
    conn.commit()
    return len(rows_to_insert)


def ingest_all(
    conn: sqlite3.Connection,
    clients_csv: Path,
    reservations_csv: Path,
    invoices_csv: Path,
    survey_csv: Path,
) -> dict[str, int]:
    """Convenience loader for synthetic CSV pipeline."""
    return {
        "clients": ingest_clients(conn, clients_csv),
        "reservations": ingest_reservations(conn, reservations_csv),
        "invoices": ingest_invoices(conn, invoices_csv),
        "surveys": ingest_survey(conn, survey_csv),
    }


# ── Real FSP XLSX ingest (incremental — preserves flight_overrides) ─────────

def _str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _float(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _parse_date(v) -> str | None:
    """Parse FSP date (M/D/YYYY string or datetime object) → YYYY-MM-DD."""
    from datetime import datetime
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s.split()[0], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # fallback: return as-is


def _load_workbook(xlsx_path: Path):
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")
    return openpyxl.load_workbook(xlsx_path, data_only=True)


# FSP doesn't currently export a per-flight rating label, but if/when it does
# (or if an Ops team adds a sidecar column), these are the header names we'll
# accept. Normalized to upper-case for case-insensitive matching.
_RATING_LABEL_HEADERS = {"RATING", "RATING LABEL", "COURSE", "PROGRAM"}

# Canonical rating codes we recognize.
_KNOWN_RATING_LABELS = {"PPL", "IFR", "COM", "AMEL", "CFI", "CFII", "MEI"}

# Map common synonyms to canonical codes.
_RATING_LABEL_SYNONYMS = {
    "ASEL COM": "COM", "COMMERCIAL": "COM", "COMMERCIAL SE": "COM",
    "PRIVATE": "PPL", "PRIVATE PILOT": "PPL",
    "INSTRUMENT": "IFR",
    "MULTI": "AMEL", "MULTI-ENGINE": "AMEL", "MULTI ENGINE": "AMEL",
}


def _normalize_rating_label(value) -> str | None:
    """Map a freeform rating tag to our canonical code, or None."""
    s = _str(value)
    if not s:
        return None
    upper = s.upper()
    if upper in _KNOWN_RATING_LABELS:
        return upper
    return _RATING_LABEL_SYNONYMS.get(upper)


def _find_rating_label_header(headers: list[str]) -> str | None:
    """Return the actual header name to read rating labels from, if any."""
    for h in headers:
        if h.upper() in _RATING_LABEL_HEADERS:
            return h
    return None


def ingest_flight_detail_xlsx(conn: sqlite3.Connection, xlsx_path: Path) -> dict[str, int]:
    """Incremental ingest of FSP Flight Detail XLSX → flights table.

    Behavior:
        - INSERT new flights (by fsp_reservation UNIQUE constraint).
        - UPDATE existing flights' FSP-sourced columns on every re-import (so
          status changes, instructor changes, etc. flow through).
        - Manual overrides in flight_overrides override these values; reapplied
          in apply_overrides() which the caller must run after ingest.

    Returns {'inserted': N, 'updated': M, 'skipped_no_resno': K}.
    """
    wb = _load_workbook(xlsx_path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in
               next(ws.iter_rows(min_row=1, max_row=1))]
    rating_header = _find_rating_label_header(headers)

    inserted = updated = skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rd = dict(zip(headers, row))

        res_num = _str(rd.get("Reservation #"))
        if not res_num:
            skipped += 1
            continue

        flight_date = _parse_date(rd.get("Date"))
        if not flight_date:
            skipped += 1
            continue

        aircraft_str = _str(rd.get("Aircraft"))
        tail, make, model = parse_aircraft_string(aircraft_str)
        ac_class = classify_aircraft(model)
        hobbs = _float(rd.get("Hobbs (Total)"))
        is_ground = 1 if (tail is None and hobbs is None) else 0
        billing_cat = derive_billing_category(rd)
        client_raw = _str(rd.get("Client")) or ""
        flight_num = _str(rd.get("Flight #"))
        res_type = _str(rd.get("Type")) or "Dual Flight Training"
        status = _str(rd.get("Status")) or "Completed"
        instructor = _str(rd.get("Instructor"))
        rating_label = (_normalize_rating_label(rd.get(rating_header))
                        if rating_header else None)

        # Does this reservation already exist?
        existing = conn.execute(
            "SELECT flight_id FROM flights WHERE fsp_reservation = ?",
            (res_num,),
        ).fetchone()

        if existing is None:
            conn.execute(
                """INSERT INTO flights (
                       fsp_reservation, fsp_flight_num, flight_date, length_hrs,
                       reservation_type, status, client_raw, aircraft_tail, aircraft_make,
                       aircraft_model, aircraft_class, instructor,
                       hobbs_hours, billing_category, is_ground_lesson, rating_label
                   ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    res_num, flight_num, flight_date, 0.0,
                    res_type, status, client_raw, tail, make,
                    model, ac_class, instructor,
                    hobbs, billing_cat, is_ground, rating_label,
                ),
            )
            inserted += 1
        else:
            # Refresh FSP-sourced fields. Overrides will re-stomp the values
            # they care about after apply_overrides() runs.
            # Preserve a previously-set rating_label only if FSP didn't supply one.
            if rating_label is None:
                conn.execute(
                    """UPDATE flights SET
                           fsp_flight_num   = ?,
                           flight_date      = ?,
                           reservation_type = ?,
                           status           = ?,
                           client_raw       = ?,
                           aircraft_tail    = ?,
                           aircraft_make    = ?,
                           aircraft_model   = ?,
                           aircraft_class   = ?,
                           instructor       = ?,
                           hobbs_hours      = ?,
                           billing_category = ?,
                           is_ground_lesson = ?
                       WHERE fsp_reservation = ?""",
                    (
                        flight_num, flight_date, res_type, status, client_raw,
                        tail, make, model, ac_class, instructor,
                        hobbs, billing_cat, is_ground,
                        res_num,
                    ),
                )
            else:
                conn.execute(
                    """UPDATE flights SET
                           fsp_flight_num   = ?,
                           flight_date      = ?,
                           reservation_type = ?,
                           status           = ?,
                           client_raw       = ?,
                           aircraft_tail    = ?,
                           aircraft_make    = ?,
                           aircraft_model   = ?,
                           aircraft_class   = ?,
                           instructor       = ?,
                           hobbs_hours      = ?,
                           billing_category = ?,
                           is_ground_lesson = ?,
                           rating_label     = ?
                       WHERE fsp_reservation = ?""",
                    (
                        flight_num, flight_date, res_type, status, client_raw,
                        tail, make, model, ac_class, instructor,
                        hobbs, billing_cat, is_ground, rating_label,
                        res_num,
                    ),
                )
            updated += 1

    conn.commit()
    return {"inserted": inserted, "updated": updated, "skipped_no_resno": skipped}


# Invoice line-item category heuristic — derived from Line Item Name.
# Names commonly seen in Provectus's invoices: "CFI Ryan Pineda", "N24342",
# "Discount", "MasterCard" / "Cash" (payments).
def _classify_invoice_line(name: str | None, line_type: str | None) -> str:
    if line_type and line_type.strip().lower() == "payment":
        return "Payment"
    if not name:
        return "Other"
    n = name.strip()
    if n.upper().startswith(("CFI ", "MEI ")) or " CFI" in n.upper():
        return "Instructor"
    # Tail numbers — start with N + digit
    if len(n) >= 2 and n[0].upper() == "N" and n[1].isdigit():
        return "Aircraft"
    if "discount" in n.lower() or "credit" in n.lower():
        return "Discount"
    if "ground" in n.lower() or "sim" in n.lower():
        return "Ground"
    return "Other"


def ingest_invoice_xlsx(conn: sqlite3.Connection, xlsx_path: Path) -> int:
    """Reload invoices table from the FSP Invoice Detail XLSX.

    Invoices have no manual-override surface (yet), so this is truncate-and-reload
    rather than incremental UPSERT. Eliminates any dedup-key headache from
    multi-line invoices that lack a per-line sequence number.

    Returns: number of rows inserted.
    """
    wb = _load_workbook(xlsx_path)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in
               next(ws.iter_rows(min_row=1, max_row=1))]

    # Pre-build reservation → flight_id map (avoids per-row SELECT in the loop)
    res_map = {
        row["fsp_reservation"]: row["flight_id"]
        for row in conn.execute("SELECT flight_id, fsp_reservation FROM flights")
    }

    rows_to_insert = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rd = dict(zip(headers, row))

        invoice_num = _str(rd.get("Invoice #"))
        if not invoice_num:
            continue

        res_num = _str(rd.get("Reservation #"))
        flight_id = res_map.get(res_num) if res_num else None

        line_type = _str(rd.get("Line Item Type"))
        line_name = _str(rd.get("Line Item Name"))
        line_desc = _str(rd.get("Line Item Description"))
        category = _classify_invoice_line(line_name, line_type)

        qty = _float(rd.get("Quantity"))
        rate = _float(rd.get("Rate"))
        cost = _float(rd.get("Total Costs"))
        payment = _float(rd.get("Total Payment"))
        # amount is signed: charges positive, payments negative
        if payment is not None and payment != 0:
            amount = -payment
        elif cost is not None:
            amount = cost
        else:
            continue  # blank money row, skip

        desc = " — ".join(x for x in [line_name, line_desc] if x) or None
        invoice_date = _parse_date(rd.get("Invoice Date")) or ""
        status = _str(rd.get("Status")) or "Unknown"

        rows_to_insert.append((
            invoice_num, invoice_date, None, flight_id,
            desc, category, qty, rate, amount, status,
        ))

    # Truncate-and-reload
    conn.execute("DELETE FROM invoices")
    conn.executemany(
        """INSERT INTO invoices (
               fsp_invoice, invoice_date, student_id, flight_id,
               description, category, qty, rate, amount, status
           ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        rows_to_insert,
    )
    conn.commit()
    return len(rows_to_insert)


# ── Auto-populate students from flight client names ──────────────────────────

def auto_populate_students_from_flights(conn: sqlite3.Connection) -> dict[str, int]:
    """Insert one student row per distinct flights.client_raw, and backfill
    flights.student_id so joins work.

    Used in the real-data path before survey responses arrive. Each generated
    student gets match_status='auto_from_flights' so reconcile.reconcile() can
    later upgrade them to 'matched' when survey data lands.

    Idempotent: re-running on a populated DB only inserts genuinely new clients
    and only backfills flights that don't already have a student_id.

    Returns {'inserted': N, 'backfilled': M}.
    """
    # Collect distinct client_raw values from flights — first non-empty token
    # before any comma (multi-client flights use comma-separated names).
    raw_rows = conn.execute(
        "SELECT DISTINCT client_raw FROM flights "
        "WHERE client_raw IS NOT NULL AND TRIM(client_raw) != ''"
    ).fetchall()

    inserted = 0
    name_to_id: dict[str, int] = {}
    for r in raw_rows:
        primary = r["client_raw"].split(",")[0].strip()
        if not primary:
            continue
        existing = conn.execute(
            "SELECT student_id FROM students WHERE fsp_display_name = ?",
            (primary,),
        ).fetchone()
        if existing is None:
            cur = conn.execute(
                """INSERT INTO students (fsp_display_name, match_status)
                   VALUES (?, 'auto_from_flights')""",
                (primary,),
            )
            name_to_id[primary] = cur.lastrowid
            inserted += 1
        else:
            name_to_id[primary] = existing["student_id"]
    conn.commit()

    # Backfill flights.student_id for any flight that doesn't have one yet.
    # Match on the primary (pre-comma) name.
    backfilled = 0
    to_update = conn.execute(
        "SELECT flight_id, client_raw FROM flights WHERE student_id IS NULL "
        "AND client_raw IS NOT NULL AND TRIM(client_raw) != ''"
    ).fetchall()
    for f in to_update:
        primary = f["client_raw"].split(",")[0].strip()
        sid = name_to_id.get(primary)
        if sid is None:
            continue
        conn.execute(
            "UPDATE flights SET student_id = ? WHERE flight_id = ?",
            (sid, f["flight_id"]),
        )
        backfilled += 1
    conn.commit()

    return {"inserted": inserted, "backfilled": backfilled}


# ── Override application ─────────────────────────────────────────────────────

def apply_overrides(conn: sqlite3.Connection) -> int:
    """Stamp flight_overrides values back onto the flights table.

    Run this AFTER every ingest pass so manual tweaks survive a re-import.
    Returns number of override rows applied.
    """
    rows = conn.execute(
        "SELECT flight_id, field_name, value FROM flight_overrides"
    ).fetchall()

    applied = 0
    for r in rows:
        # Only allow whitelisted columns to be overridden — never let the table
        # become an arbitrary SQL injection surface.
        col = r["field_name"]
        if col not in _OVERRIDABLE_COLUMNS:
            continue
        raw = r["value"]
        cast = _OVERRIDABLE_COLUMNS[col](raw) if raw is not None else None
        conn.execute(
            f"UPDATE flights SET {col} = ? WHERE flight_id = ?",
            (cast, r["flight_id"]),
        )
        applied += 1
    conn.commit()
    return applied


def _to_int(s: str) -> int:
    return int(s)


def _to_str(s: str) -> str:
    return str(s)


# Whitelist of columns on the flights table that can be overridden via UI.
# Maps column name → caster from stored TEXT value back to typed value.
_OVERRIDABLE_COLUMNS = {
    "is_ground_lesson": _to_int,
    "billing_category": _to_str,
    "aircraft_class":   _to_str,
    "reservation_type": _to_str,
    "rating_label":     _to_str,   # manual per-flight rating attribution
}


def set_flight_override(
    conn: sqlite3.Connection,
    flight_id: int,
    field_name: str,
    value,
    note: str | None = None,
) -> None:
    """Upsert a single override. `value` can be any type; stored as TEXT."""
    if field_name not in _OVERRIDABLE_COLUMNS:
        raise ValueError(f"Field {field_name!r} is not overridable. "
                         f"Allowed: {sorted(_OVERRIDABLE_COLUMNS)}")
    from datetime import datetime
    conn.execute(
        """INSERT INTO flight_overrides (flight_id, field_name, value, note, set_at)
           VALUES (?, ?, ?, ?, ?)
           ON CONFLICT (flight_id, field_name) DO UPDATE SET
               value = excluded.value,
               note  = excluded.note,
               set_at = excluded.set_at""",
        (flight_id, field_name, str(value) if value is not None else None,
         note, datetime.utcnow().isoformat(timespec="seconds")),
    )
    conn.commit()


def clear_flight_override(conn: sqlite3.Connection, flight_id: int,
                          field_name: str) -> None:
    conn.execute(
        "DELETE FROM flight_overrides WHERE flight_id = ? AND field_name = ?",
        (flight_id, field_name),
    )
    conn.commit()
